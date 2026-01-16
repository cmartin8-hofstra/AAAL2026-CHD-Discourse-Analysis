#!/usr/bin/env python3
"""
Script 06: Multimodal Coding Template Generator (Gill & Lennon, 2022)

Generates structured Excel template for manual qualitative coding of visual
meta-functions in selected exemplar tweets.

Theoretical Framework:
    - Gill, P., & Lennon, R. (2022). A multimodal fear appeals analysis
      framework for vaccine misinformation. Journal of Medical Internet
      Research, 24(6), e36255.

      Gill & Lennon's framework analyzes visual fear appeals across five
      meta-functions:
      1. Composition - Visual arrangement and design elements
      2. Color - Affective impact through color choices
      3. Represented Participants (RP) - Who/what is depicted
      4. Perspective/Angles - Camera position and viewer relationship
      5. Textual Components - Text overlays, captions, embedded language

    - Androutsopoulos, J. (2014). Mediatization and sociolinguistic change.
      De Gruyter.

      Digital discourse analysis emphasizes multimodal integration of
      linguistic and visual semiotic resources.

Template Structure:
    - 15 sheets (one per CHD state chapter)
    - 10 rows per sheet (top 10 exemplars from Script 05)
    - Column A: Tweet #/Identifier (e.g., MAIN_16.png)
    - Columns B-G: Coding categories (initially blank for manual coding)

Output Files:
    - data/analyzed/multimodal_coding_template.xlsx (blank template)
    - data/analyzed/coding_template_report.txt (instructions and metadata)

Reproducibility Note (APA 7th):
    Template pre-populated with exemplar IDs maintains traceability to
    computational pipeline (Scripts 01-05). Blank coding fields ensure
    qualitative analysis independence from automated features.

Author: AAAL 2026 Pipeline
Date: January 2026
"""
import os
import sys
from datetime import datetime
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Project paths
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
os.chdir(PROJECT_ROOT)


def load_config():
    """Load configuration from config.yaml with error handling (APA reproducibility)."""
    config_path = os.path.join(PROJECT_ROOT, 'config', 'config.yaml')
    try:
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
        return config
    except Exception as e:
        raise ValueError(f"Config loading error: {e}")


def setup_logging(config):
    """Setup log file with timestamp."""
    log_dir = os.path.join(PROJECT_ROOT, config['paths']['log_dir'])
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_path = os.path.join(log_dir, f'06_multimodal_coding_template_{timestamp}.log')
    return log_path, timestamp


def log_message(log_path, message, also_print=True):
    """Write message to log file and optionally print."""
    with open(log_path, 'a') as f:
        f.write(f"{datetime.now().strftime('%H:%M:%S')} | {message}\n")
    if also_print:
        print(message)


def create_coding_template(exemplars_df, config, log_path):
    """
    Create Excel coding template with styling (APA 7th table compliance).
    """
    wb = Workbook()
    stats = {'sheets_created': 0, 'total_rows': 0}

    # Coding categories (Gill & Lennon, 2022 framework)
    categories = [
        'Composition (arrangement, design)',
        'Color (affective impact)',
        'Represented Participants (who/what depicted)',
        'Perspective/Angles (viewer relationship)',
        'Textual Components (overlays, embedded text)',
        'Fear Appeal Integration (overall strategy)'
    ]

    # Styling for APA compliance (clear, readable, no gridlines)
    header_font = Font(bold=True, size=12)
    cell_font = Font(size=11)
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for chapter_full in config['chapters']:
        chapter_code = chapter_full.replace('CHD_', '').rstrip('_')
        chapter_exemplars = exemplars_df[exemplars_df['chapter_code'] == chapter_code]

        if len(chapter_exemplars) == 0:
            continue

        ws = wb.create_sheet(title=chapter_full)
        stats['sheets_created'] += 1

        # Headers
        ws['A1'] = 'Tweet ID'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = alignment
        ws['A1'].border = thin_border

        for col, cat in enumerate(categories, start=2):
            cell = ws.cell(row=1, column=col, value=cat)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # Populate exemplar IDs (e.g., MAIN_16.png)
        for row_num, (_, exemplar) in enumerate(chapter_exemplars.iterrows(), start=2):
            ws.cell(row=row_num, column=1, value=f"{exemplar['data_snap_source']}.png")
            stats['total_rows'] += 1

            # Style data cells
            for col in range(1, len(categories) + 2):
                cell = ws.cell(row=row_num, column=col)
                cell.font = cell_font
                cell.alignment = alignment
                cell.border = thin_border

        # Adjust column widths (APA readability)
        ws.column_dimensions['A'].width = 15
        for col_letter in 'BCDEF':
            ws.column_dimensions[col_letter].width = 25

        log_message(log_path, f"  Sheet {chapter_full}: {len(chapter_exemplars)} rows")

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output_path = os.path.join(PROJECT_ROOT, config['paths']['analyzed_dir'], 'multimodal_coding_template.xlsx')
    wb.save(output_path)

    log_message(log_path, f"  Template saved: {output_path} ({stats['total_rows']} rows)")
    log_message(log_path, "")

    return stats, output_path


def main():
    """Main execution function."""
    print("=" * 60)
    print("SCRIPT 06: MULTIMODAL CODING TEMPLATE")
    print("=" * 60)
    print()

    config = load_config()
    log_path, timestamp = setup_logging(config)

    log_message(log_path, "=" * 60)
    log_message(log_path, "AAAL 2026 MULTIMODAL CODING TEMPLATE")
    log_message(log_path, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log_message(log_path, "=" * 60)
    log_message(log_path, "")

    # =========================================================================
    # Step 1: Load Exemplars Data (CSV from Script 05)
    # =========================================================================
    log_message(log_path, "STEP 1: LOADING EXEMPLARS DATA")
    log_message(log_path, "-" * 40)

    input_path = os.path.join(PROJECT_ROOT, config['paths']['analyzed_dir'], 'selected_exemplars.csv')

    if not os.path.exists(input_path):
        log_message(log_path, f"ERROR: Input file missing: {input_path}")
        return 1

    exemplars_df = pd.read_csv(input_path, low_memory=False)

    log_message(log_path, f"  Loaded {len(exemplars_df)} exemplars from {input_path}")
    log_message(log_path, "")

    # =========================================================================
    # Step 2: Generate Coding Template
    # =========================================================================
    log_message(log_path, "STEP 2: GENERATING CODING TEMPLATE")
    log_message(log_path, "-" * 40)

    stats, output_path = create_coding_template(exemplars_df, config, log_path)

    # =========================================================================
    # Step 3: Generate Coding Template Report
    # =========================================================================
    log_message(log_path, "STEP 3: GENERATING TEMPLATE REPORT")
    log_message(log_path, "-" * 40)

    report_path = os.path.join(PROJECT_ROOT, config['paths']['analyzed_dir'], 'coding_template_report.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("AAAL 2026 MULTIMODAL CODING TEMPLATE REPORT\n")
        f.write(f"Generated: {timestamp}\n")
        f.write("=" * 60 + "\n\n")

        f.write("THEORETICAL FRAMEWORK:\n")
        f.write("  Gill, P., & Lennon, R. (2022). A multimodal fear appeals analysis\n")
        f.write("  framework for vaccine misinformation. JMIR, 24(6), e36255.\n\n")
        f.write("  Androutsopoulos, J. (2014). Mediatization and sociolinguistic change.\n")
        f.write("  De Gruyter.\n\n")

        f.write("CODING INSTRUCTIONS (Gill & Lennon, 2022):\n")
        f.write("  1. Composition: Describe visual arrangement, salience, framing.\n")
        f.write("  2. Color: Note palette choices and emotional/affective impact.\n")
        f.write("  3. Represented Participants: Identify depicted actors/objects and roles.\n")
        f.write("  4. Perspective/Angles: Analyze viewer positioning (e.g., power relations).\n")
        f.write("  5. Textual Components: Examine embedded text, typography, integration.\n")
        f.write("  6. Fear Appeal: Synthesize how elements construct threat/vulnerability.\n\n")

        f.write("TEMPLATE SUMMARY:\n")
        f.write(f"  Sheets created: {stats['sheets_created']}\n")
        f.write(f"  Total coding rows: {stats['total_rows']} (target: 150)\n\n")

        f.write("CHAPTER BREAKDOWN:\n")
        f.write(f"  {'Chapter':<15} {'Exemplars':>10}\n")
        f.write(f"  {'-'*15} {'-'*10}\n")
        chapter_exemplars = exemplars_df.groupby('chapter_code').size().to_dict()
        for chapter_code in sorted(chapter_exemplars.keys()):
            f.write(f"  {chapter_code:<15} {chapter_exemplars[chapter_code]:>10}\n")

        f.write("\nPIPELINE TRACEABILITY:\n")
        f.write("  Script 01 → Data standardization (1:1 multimodal alignment)\n")
        f.write("  Script 02 → PMI lexical association analysis\n")
        f.write("  Script 03 → Epistemic stance classification (Heritage, 2012)\n")
        f.write("  Script 04 → Sentiment analysis & corpus validation\n")
        f.write("  Script 05 → Exemplar extraction (composite scoring)\n")
        f.write("  Script 06 → THIS TEMPLATE (manual qualitative coding)\n\n")

        f.write("METHODOLOGICAL INTEGRATION:\n")
        f.write("  Computational analysis (Scripts 01-05) identifies prototypical\n")
        f.write("  instances using PMI, epistemic density, and sentiment intensity.\n")
        f.write("  Manual qualitative coding (Script 06) examines visual fear appeal\n")
        f.write("  strategies following Gill & Lennon (2022) framework.\n\n")

        f.write("OUTPUT FILES:\n")
        f.write(f"  {output_path}\n")
        f.write(f"    -> Blank coding template, {stats['total_rows']} rows\n")
        f.write(f"  outputs/exemplars/{{chapter}}/\n")
        f.write(f"    -> PNG files for visual reference during coding\n")

    log_message(log_path, f"  Report saved: {report_path}")
    log_message(log_path, "")

    # =========================================================================
    # Final Summary
    # =========================================================================
    log_message(log_path, "=" * 60)
    log_message(log_path, "SCRIPT 06 COMPLETE - CODING TEMPLATE GENERATED")
    log_message(log_path, "=" * 60)
    log_message(log_path, f"Template created: {output_path}")
    log_message(log_path, f"Sheets: {stats['sheets_created']}")
    log_message(log_path, f"Coding rows: {stats['total_rows']}")
    log_message(log_path, f"Instructions: {report_path}")
    log_message(log_path, "")
    log_message(log_path, "NEXT STEPS:")
    log_message(log_path, "  1. Open multimodal_coding_template.xlsx")
    log_message(log_path, "  2. Reference PNGs in outputs/exemplars/{chapter}/")
    log_message(log_path, "  3. Code each exemplar using Gill & Lennon (2022) framework")
    log_message(log_path, "  4. Save completed coding for qualitative analysis")
    log_message(log_path, f"Log: {log_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
